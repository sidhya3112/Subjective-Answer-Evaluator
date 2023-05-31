import openpyxl
import utils
import spacy
from nltk.corpus import stopwords
import pytextrank
import utils
import yake
import en_core_web_sm

# Load the spaCy model
nlp = spacy.load("en_core_web_sm")

# add PyTextRank to the spaCy pipeline
nlp.add_pipe("textrank")


def ans_key(sheet):
    utils.finalList.clear()
    utils.all_keywords.clear()
    utils.list_of_strings.clear()
    
    utils.finalList = get_anslst(sheet)
    utils.list_of_strings = print_ans(utils.finalList)
    utils.all_keywords = get_keywords(utils.finalList)

	

def get_anslst(sheet):
    finalList = []
    startRow = 7
    endRow = 9
    maxRow = sheet.max_row 
    # print(maxRow)
	#cell = sheet.cell(row= startRow, column = 1)
    while endRow <= maxRow:
        anslst = []
        for j in range(startRow, endRow+1):
            data = sheet.cell(row = j, column = 1)
            anslst.append(data.value)
            # print(anslst)

		#print(len(anslst))
        finalList.append(anslst) 
        # print(finalList)   
        startRow+=6
        endRow+= 6	
        

    # print("Final list is :")
    # print(finalList)
    # print(len(finalList))

    return finalList


def get_keywords(finalList):
    #Keyword Extraction:

   
    all_keywords = []
    stop_words = set(stopwords.words('english'))
    keyword_extractor = yake.KeywordExtractor(top=10, stopwords=stop_words)
    
    for ans in range(len(finalList)): 
        keyword = []
        templist = finalList[ans]
        for sent in templist:
            keywords = keyword_extractor.extract_keywords(sent)
            top_keywords = [key[0] for key in keywords]
            keyword = keyword + top_keywords
        keyword = set(keyword)
        keyword = list(keyword)
        all_keywords.append(keyword)
	
    return all_keywords
	


#Print model answer key

def print_ans(finalList):
    list_of_strings = []
   
    for i in range(len(finalList)):
        sample_ans = ""
        temp = finalList[i]
        for j in range(len(temp)):
            sample_ans = temp[j]
          #sample_ans + "\n\n" + str(j+1) + ") " +
        list_of_strings.append(sample_ans)
	# print(sample_ans)
    return list_of_strings

   