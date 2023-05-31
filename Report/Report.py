import openpyxl
import tkinter as tk
from tkinter import *
from tkinter import messagebox
from Report.get_keyword_score import get_keyword_score
from Report.get_similarity_score import get_similarity_score
from Report.get_grammar_score import get_grammar_score
from Report.get_spelling_score import get_spelling_score
import utils
from Report.Det_Report import Det_Report


class Report(tk.Frame):

		#new.geometry("100x50+665+410")
	def __init__(self):
		new =tk.Frame.__init__(self)
		new = Toplevel(self)
		#new.pack()
		new.geometry("350x180+500+300")
		new.title("Evaluation Report")

		get_similarity_score(utils.list_user_input, utils.finalList)
		get_keyword_score(utils.list_user_input, utils.all_keywords)
		get_spelling_score(utils.list_user_input)
		get_grammar_score(utils.list_user_input)



		similarity_factor = 0.4
		keyword_factor = 0.3
		grammar_factor = 0.15
		spell_factor = 0.15

		Qtext = utils.Qtext
		result = ['']* len(Qtext)
		all_final_score = ['']* len(Qtext)
		
		total = 0
		print(len(utils.all_spelling_scores))
		print(len(utils.all_grammar_scores))
		print(len(utils.all_similarity_scores))
		

		
		#utils.all_keyword_scores*keyword_factor

		for i in range(len(Qtext)):
			result[i] = utils.all_similarity_scores[i]*similarity_factor+ utils.all_keyword_scores[i]*keyword_factor+ utils.all_grammar_scores[i]*grammar_factor + utils.all_spelling_scores[i]*spell_factor
			all_final_score[i] = round(result[i], 2)
			if(utils.all_keyword_scores[i]<=0.2):
				all_final_score[i] = 0
			 


		total = sum(all_final_score)*10
		total = round(total,2)
		utils.total = total
		

		utils.all_final_score = all_final_score
		op = "users.xlsx"
		workbook = openpyxl.load_workbook(op)
		sheet = workbook.active
		next_row = sheet.max_row + 1
		r=2
		for i in range(2, next_row):
			if sheet.cell(row=i, column=1).value == utils.user:
				r = i
				break
		sheet.cell(row=r, column = get_workbook_column()).value = total
		workbook.save(op)
		
		
		new.label = tk.Label(new, text=" Your Marks is = " + str(total)+ " out of 100")
		#new.label.pack(side='left')
		new.label.grid(row= 2, column=2, padx=10, pady=15,sticky=N)
		new.button = tk.Button( new, text = "Detailed Report", width = 15,command = self.Det_report )
		#new.button.pack(side='right')
		new.button.grid(row= 3, column=3, padx=10, pady=15,sticky=N)
		new.button2 = tk.Button( new, text = "Close", width = 15,command = self.close_window )
		#new.button.pack(side='right')
		new.button2.grid(row= 4, column=3, padx=10, pady=15,sticky=N)
	def Det_report(self):
		self.newWindow = Det_Report()
		# self.hide()
	
	def close_window(self):
		self.master.destroy()

def get_workbook_column():
		col=3
		if utils.file_name == "Cryptography":
			col = 3
		elif utils.file_name == "Cyber-Security":
			col = 4
		elif utils.file_name == "E-Commerce":
			col = 5
		elif utils.file_name == "NLP":
			col = 6
		elif utils.file_name == "Philosophy":
			col = 7

		return col

