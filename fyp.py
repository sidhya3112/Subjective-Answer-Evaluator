import tkinter as tk
from tkinter import *
import openpyxl
from Test import Test
from openmyfile import openmyfile
import utils
from PIL import ImageTk, Image
from tkinter import ttk

from tkinter import Tk, Toplevel, StringVar, Label, Entry, Frame, OptionMenu, Button, messagebox


class SignUp(tk.Toplevel):
    def __init__(self, master=None):
        tk.Toplevel.__init__(self, master)
        self.title("Sign Up")
        self.resizable(False, False)
        self.tk_setPalette(background='#fff')
        self.geometry("925x500+300+200")


        # Add the image
        img_path = 'signup.png'
        self.display_image(img_path, 75, 50)

        frame = Frame(self, width=350, height=350, bg="white")
        frame.place(x=480, y=70)
        
        heading = tk.Label(frame, text='Sign Up', fg="#57a1f8", bg='white', font=('Comic Sans MS', 30, 'bold'))
        heading.place(x=100, y=5)


        # username
        def on_leave_username(e):
            name = self.user.get()
            if name == '':
                self.user.delete(0, tk.END)
                self.user.insert(0, 'Username')

        def on_enter_username(e):
            self.user.delete(0, tk.END)

        self.user = Entry(frame, width=25, fg='black', border=0, bg='white', font=('Microsoft YaHei UI Light', 11))
        self.user.place(x=30, y=80)
        self.user.insert(0, 'Username')
        self.user.bind('<FocusIn>', on_enter_username)
        self.user.bind('<FocusOut>', on_leave_username)

        Frame(frame, width=295, height=2, bg='black').place(x=25, y=107)

        # passcode
        def on_leave_passcode(e):
            name = self.passcode.get()
            if name == '':
                self.passcode.delete(0, tk.END)
                self.passcode.insert(0, 'Password')

        def on_enter_passcode(e):
            self.passcode.delete(0, tk.END)

        self.passcode = Entry(frame, width=25, fg='black', border=0, bg='white', font=('Microsoft YaHei UI Light', 11))
        self.passcode.place(x=30, y=130)
        self.passcode.insert(0, 'Password')
        self.passcode.bind('<FocusIn>', on_enter_passcode)
        self.passcode.bind('<FocusOut>', on_leave_passcode)

        Frame(frame, width=295, height=2, bg='black').place(x=25, y=157)

        # confirm passcode
        def on_leave_conf_passcode(e):
            name = self.conf_passcode.get()
            if name == '':
                self.conf_passcode.delete(0, tk.END)
                self.conf_passcode.insert(0, 'Confirm Password')

        def on_enter_conf_passcode(e):
            self.conf_passcode.delete(0, tk.END)

        self.conf_passcode = Entry(frame, width=25, fg='black', border=0, bg='white', font=('Microsoft YaHei UI Light', 11))
        self.conf_passcode.place(x=30, y=180)
        self.conf_passcode.insert(0, 'Confirm Password')
        self.conf_passcode.bind('<FocusIn>', on_enter_conf_passcode)
        self.conf_passcode.bind('<FocusOut>', on_leave_conf_passcode)

        Frame(frame, width=295, height=2, bg='black').place(x=25, y=207)

        Button(frame, width=39, pady=7, text='Sign Up', bg='#57a1f8', fg='white', border=0, command=self.click_submit).place(x=35, y=234)

        label = Label(frame, text="Existing account? ", fg="black", bg="white", font=('Microsoft YaHei UI Light', 10, 'bold'))
        label.place(x=75, y=280)

        login = Button(frame, width=6, text="Sign In", border=0, bg='white', cursor='hand2', fg="#57a1f8",
                       font=('Microsoft YaHei UI Light', 9, 'bold'), command=self.click_signin)
        login.place(x=245, y=280)

    def click_submit(self):
        if self.passcode.get() == self.conf_passcode.get():
            user = self.user.get()
            password = self.passcode.get()
        else:
            messagebox.showinfo("Signup error", "Password and confirm password do not match!")
            return

        op = "users.xlsx"
        workbook = openpyxl.load_workbook(op)
        sheet = workbook.active
        next_row = sheet.max_row + 1
        for i in range(2, next_row):
            if sheet.cell(row=i, column=1).value == user:
                messagebox.showinfo("Signup error", "User already exists!")
                break
        else:
            sheet.cell(row=next_row, column=1).value = user
            sheet.cell(row=next_row, column=2).value = password
            workbook.save(op)
            utils.user = user

            # #answers.xlsx
            # workbook = openpyxl.load_workbook("answers.xlsx")
            # sheet = workbook.active
            # next_row = sheet.max_row + 1
            # sheet.cell(row=next_row, column=1).value = user
            # workbook.save(op)

            self.newWindow = Test()
            self.destroy()
            root.withdraw()

    def display_image(self, img_path, x, y):
        img = Image.open(img_path)
        img = img.resize((300, 300), Image.LANCZOS)
        img = ImageTk.PhotoImage(img)
        img_label = Label(self, image=img, bg='white')
        img_label.image = img  # Keep a reference to prevent garbage collection
        img_label.place(x=x, y=y)

    def click_signin(self):
        self.destroy()  # Destroy the sign-up window
        root.deiconify()  # Show the existing sign-in window



class App(tk.Frame):
    def __init__(self):
        tk.Frame.__init__(self)
        self.pack()

        self.master.tk_setPalette(background='white')
        self.master.geometry('925x500+300+200')
        self.master.title("Sign In")

        # Create a Tkinter variable
        tkt = StringVar(root)

        global img

        # Add the image
        img_path = 'login.png'
        self.display_image(img_path, 75, 50)

        frame = Frame(self.master, width=350, height=350, bg="white")
        frame.place(x=480, y=70)

        heading = Label(frame, text='Sign In', fg="#57a1f8", bg='white', font=('Comic Sans MS', 30, 'bold'))
        heading.place(x=100, y=5)

        # username
        def on_leave_username(e):
            name = self.user.get()
            if name == '':
                self.user.delete(0, tk.END)
                self.user.insert(0, 'Username')

        def on_enter_username(e):
            self.user.delete(0, tk.END)

        self.user = Entry(frame, width=25, fg='black', border=0, bg='white', font=('Microsoft YaHei UI Light', 11))
        self.user.place(x=30, y=80)
        self.user.insert(0, 'Username')
        self.user.bind('<FocusIn>', on_enter_username)
        self.user.bind('<FocusOut>', on_leave_username)

        Frame(frame, width=295, height=2, bg='black').place(x=25, y=107)

        # passcode
        def on_leave_passcode(e):
            name = self.passcode.get()
            if name == '':
                self.passcode.delete(0, tk.END)
                self.passcode.insert(0, 'Password')

        def on_enter_passcode(e):
            self.passcode.delete(0, tk.END)

        self.passcode = Entry(frame, width=25, fg='black', border=0, bg='white', font=('Microsoft YaHei UI Light', 11))
        self.passcode.place(x=30, y=130)
        self.passcode.insert(0, 'Password')
        self.passcode.bind('<FocusIn>', on_enter_passcode)
        self.passcode.bind('<FocusOut>', on_leave_passcode)

        Frame(frame, width=295, height=2, bg='black').place(x=25, y=157)
        # Dictionary with options
	
        # choices = {'E-Commerce', 'NLP', 'Cryptography', 'Cyber-Security', 'Philosophy'}
        # tkt.set('Cyber-Security')  # set the default option
        # openmyfile('Cyber-Security')
        # popupMenu = OptionMenu(frame, tkt, *choices)
        # subj = Label(frame, text="Subject     : ", font=('Microsoft YaHei UI Light', 11))
        # subj.place(x=30, y=180)
        # popupMenu.place(x=140, y=180)
        Button(frame, width=39, pady=7, text='Sign In', bg='#57a1f8', fg='white', border=0,
               command=self.click_ok).place(x=35, y=234)

        label = Label(frame, text="Don't have an account?", fg="black", bg="white",
                      font=('Microsoft YaHei UI Light', 10, 'bold'))
        label.place(x=75, y=280)

        sign_up = Button(frame, width=6, text="Sign Up", border=0, bg='white', cursor='hand2', fg="#57a1f8",
                         font=('Microsoft YaHei UI Light', 9, 'bold'), command=self.click_signup)
        sign_up.place(x=245, y=280)
        
        combo = ttk.Combobox(frame, textvariable = tkt, values= ['E-Commerce', 'NLP', 'Cryptography', 'Cyber-Security', 'Philosophy'])
        combo.set('Cyber-Security')
        openmyfile('Cyber-security')
        subj = Label(frame, text="Choose subject     : ", font=('Microsoft YaHei UI Light', 11, ), bg='white')
        subj.place(x=30, y=180)
        combo.place(x=180, y=180)

        def change_dropdown(*args):
            utils.file_name = tkt.get()
            openmyfile(utils.file_name)

        # link function to change dropdown
        tkt.trace('w', change_dropdown)

    def click_ok(self):
        user = self.user.get()
        password = self.passcode.get()

        op = "users.xlsx"
        workbook = openpyxl.load_workbook(op)
        sheet = workbook.active
        next_row = sheet.max_row + 1
        found=False
        for i in range(2, next_row):
            if sheet.cell(row=i, column=1).value == user and sheet.cell(row=i, column=2).value == password:
                found=True
                utils.user = user
                self.newWindow = Test()
                root.withdraw()
                break
            elif sheet.cell(row=i, column=1).value == user and sheet.cell(row=i, column=2).value != password:
                found=True
                messagebox.showinfo("Sign In error", "Incorrect password!")

        if found==False:
            messagebox.showinfo("Sign In error", "User does not Exist! Please SignUp!")


    def display_image(self, img_path, x, y):
        img = Image.open(img_path)
        img = img.resize((300, 300), Image.LANCZOS)
        img = ImageTk.PhotoImage(img)
        img_label = Label(self.master, image=img, bg='white')
        img_label.image = img  # Keep a reference to prevent garbage collection
        img_label.place(x=x, y=y)

    def click_signup(self):
        self.newWindow = SignUp()
        root.withdraw()


if __name__=='__main__':
    root = Tk()
    app = App()
    root.mainloop()