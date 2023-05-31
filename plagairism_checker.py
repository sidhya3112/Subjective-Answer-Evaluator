# Plagairism Check

import openpyxl
from sklearn.feature_extraction.text import TfidfVectorizer
from sklearn.metrics.pairwise import cosine_similarity

def get_similarity_score(answer1, answer2):

	# Create TF-IDF vectorizer
	vectorizer = TfidfVectorizer()

	# Fit and transform the documents
	tfidf_matrix = vectorizer.fit_transform([answer1, answer2])

	# Calculate cosine similarity
	similarity_matrix = cosine_similarity(tfidf_matrix[0], tfidf_matrix[1])

	# Extract the similarity score
	score = similarity_matrix[0][0]

	return score


op = "answers.xlsx"
workbook = openpyxl.load_workbook(op)
sheet = workbook.active
next_col = sheet.max_column + 1

plagairism_scores = []

for r in range(2, sheet.max_row + 1):
    user = sheet.cell(row=r, column=1).value
    question = []

    for c in range(2, next_col):
        answer1 = sheet.cell(row=r, column=c).value
        plagiarism_percent = []

        for check in range(2, sheet.max_row + 1):
            if(r==check):
                 plagiarism_percent.append(-1)
                 continue
            answer2 = sheet.cell(row=check, column=c).value
            score = get_similarity_score(answer1, answer2)
            score = round(score, 2)
            plagiarism_percent.append(score)

        question.append(plagiarism_percent)

    plagairism_scores.append([user, question])

print(plagairism_scores)


for user in plagairism_scores:
    # print(user)
    for ques_index, ques in enumerate(user[1]):
        # print(ques)
        for value_index, value in enumerate(ques):
            if value > 0.7:
                copied_user = sheet.cell(row=2 + value_index, column=1).value
                print(user[0] + " has copied question " + str(ques_index+1) +" from " + copied_user)
